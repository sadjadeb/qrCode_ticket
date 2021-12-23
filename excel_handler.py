import openpyxl
import random


def get_users_from_excel(file_path):
    """
    Get users data from excel file
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    users_list = []
    for i in range(1, ws.max_row + 1):
        if i == 1:
            if ws.cell(row=1, column=ws.max_column).value != 'ticket_id':
                raise Exception('Excel file is not correct. Please provide a ticket id for users. You can add it simply by running the add_ticket_id command.')
        if i > 1:
            users_list.append({
                'first_name': ws.cell(row=i, column=1).value,
                'last_name': ws.cell(row=i, column=2).value,
                'phone_number': ws.cell(row=i, column=3).value,
                'ticket_id': ws.cell(row=i, column=ws.max_column).value,
            })

    return users_list


def add_ticket_id(file_path):
    """
    Add ticket_id to users
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws_max_column = ws.max_column
    ws.cell(row=1, column=ws_max_column + 1).value = 'ticket_id'

    generated_ids = []
    for i in range(2, ws.max_row + 1):
        ticket_id = random.randint(100000, 999999)
        while ticket_id in generated_ids:
            ticket_id = random.randint(100000, 999999)
        generated_ids.append(ticket_id)

        ws.cell(row=i, column=ws_max_column + 1).value = ticket_id

    wb.save(file_path)
