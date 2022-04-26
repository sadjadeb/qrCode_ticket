import openpyxl
import random


def get_users_from_excel(file_path):
    """
    Get users data from excel file
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    users_list = []
    for i in range(1, ws.max_row + 1):
        if i == 1:
            if ws.cell(row=1, column=4).value != 'ticket_id':
                raise Exception(
                    'Excel file is not correct. Please provide a ticket id for users. You can add it simply by running the generate_ticket_link command.')
        if i > 1:
            users_list.append({
                'first_name': ws.cell(row=i, column=1).value,
                'last_name': ws.cell(row=i, column=2).value,
                'phone_number': ws.cell(row=i, column=3).value,
                'ticket_id': ws.cell(row=i, column=4).value,
            })

    return users_list


def generate_ticket_link(input_file_path, output_file_path, base_url):
    """
    Create tickets link
    """
    output_wb = openpyxl.load_workbook(output_file_path)
    output_ws = output_wb.active
    output_ws.cell(row=1, column=1).value = 'first_name'
    output_ws.cell(row=1, column=2).value = 'last_name'
    output_ws.cell(row=1, column=3).value = 'phone_number'
    output_ws.cell(row=1, column=4).value = 'ticket_id'
    output_ws.cell(row=1, column=5).value = 'ticket_link'

    input_wb = openpyxl.load_workbook(input_file_path)
    input_ws = input_wb.active

    generated_ids = []
    for i in range(2, input_ws.max_row + 1):
        # Generate ticket id
        ticket_id = random.randint(100000, 999999)
        while ticket_id in generated_ids:
            ticket_id = random.randint(100000, 999999)
        generated_ids.append(ticket_id)

        # Store user's data, ticket id, ticket link
        output_ws.cell(row=i, column=1).value = input_ws.cell(row=i, column=1).value
        output_ws.cell(row=i, column=2).value = input_ws.cell(row=i, column=2).value
        output_ws.cell(row=i, column=3).value = input_ws.cell(row=i, column=3).value
        output_ws.cell(row=i, column=4).value = ticket_id
        output_ws.cell(row=i, column=5).value = base_url + '/ticket/' + str(ticket_id)

    output_wb.save(output_file_path)
