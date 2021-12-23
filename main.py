import openpyxl
import random
from fastapi import FastAPI
import uvicorn

app = FastAPI()


@app.get("/")
async def read_items(ticket_id: int):
    return {"ticket_id": ticket_id}


def run():
    uvicorn.run(app, host="0.0.0.0", port=8000)


def parse_users(file_path):
    """
    Parse names from a file
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    users_list = []
    for i in range(2, ws.max_row + 1):
        users_list.append({
            'first_name': ws.cell(row=i, column=1).value,
            'last_name': ws.cell(row=i, column=2).value,
            'phone_number': ws.cell(row=i, column=3).value,
            'enter_code': random.randint(10000, 99999),
        })

    return users_list


if __name__ == '__main__':
    # users = parse_users(INPUT_FILE_PATH)
    #
    # for user in users:
    #     print(user)

    run()
